"""
Excel / CSV Parser Service
Dynamically detects columns and maps to lead fields
"""
import io
from typing import List, Dict, Any, Tuple, Optional


def detect_columns(content: bytes, filename: str) -> Tuple[List[str], Dict[str, Any]]:
    """
    Detect column names from Excel or CSV file.
    Returns (column_names, sample_first_row).
    """
    if filename.lower().endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        columns = reader.fieldnames or []
        sample = dict(rows[0]) if rows else {}
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], {}
        columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
        sample = dict(zip(columns, rows[1])) if len(rows) > 1 else {}
        wb.close()

    # Clean column names
    columns = [str(c).strip() for c in columns if c]
    return columns, sample


def _normalize_header(h: str) -> str:
    """Lowercase, strip, remove special chars for fuzzy matching."""
    return h.lower().strip().replace(" ", "_").replace("-", "_")


# Auto-detection synonyms
_AUTO_MAP = {
    "email": ["email", "email_address", "e-mail", "mail"],
    "first_name": ["first_name", "firstname", "first", "given_name", "givenname"],
    "last_name": ["last_name", "lastname", "last", "surname", "family_name"],
    "company_name": ["company_name", "company", "organization", "org", "firm"],
    "company_description": ["company_description", "description", "about", "company_info", "about_company"],
    "job_title": ["job_title", "title", "position", "role", "designation", "decision_maker"],
    "linkedin_url": ["linkedin_url", "linkedin", "li_url", "linkedin_profile"],
    "why_fit": ["why_fit", "why_good_fit", "fit", "reason", "notes", "why"],
}


def _auto_detect_mapping(columns: List[str], saved_mapping: Optional[Dict] = None) -> Dict[str, str]:
    """Try to auto-map Excel columns to lead fields."""
    mapping: Dict[str, str] = {}

    # If user saved a mapping, use it as base
    if saved_mapping:
        for lead_field, excel_col in saved_mapping.items():
            if excel_col in columns:
                mapping[lead_field] = excel_col

    # Fill missing fields via fuzzy match
    for lead_field, synonyms in _AUTO_MAP.items():
        if lead_field in mapping:
            continue
        for col in columns:
            norm = _normalize_header(col)
            if norm in synonyms:
                mapping[lead_field] = col
                break

    return mapping


def parse_excel(content: bytes, filename: str, saved_mapping: Optional[Dict] = None) -> List[Dict[str, Any]]:
    """
    Parse Excel/CSV and return list of lead dicts.
    Uses saved_mapping if available, falls back to auto-detection.
    """
    if filename.lower().endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        raw_rows = [dict(row) for row in reader]
        columns = reader.fieldnames or list(raw_rows[0].keys()) if raw_rows else []
    else:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return []
        columns = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(all_rows[0])]
        raw_rows = [dict(zip(columns, row)) for row in all_rows[1:]]

    if not raw_rows:
        return []

    mapping = _auto_detect_mapping(columns, saved_mapping)
    leads = []

    for row in raw_rows:
        if not any(row.values()):
            continue  # skip empty rows

        lead: Dict[str, Any] = {}
        for lead_field, excel_col in mapping.items():
            val = row.get(excel_col)
            lead[lead_field] = str(val).strip() if val is not None else None

        # Collect unmapped columns as custom_fields
        mapped_excel_cols = set(mapping.values())
        custom: Dict[str, Any] = {}
        for col in columns:
            if col not in mapped_excel_cols:
                val = row.get(col)
                if val is not None and str(val).strip():
                    custom[col] = str(val).strip()
        lead["custom_fields"] = custom if custom else None

        # Only include rows that have at least an email
        if lead.get("email"):
            leads.append(lead)

    return leads
